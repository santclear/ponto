# -*- coding: utf-8 -*-
"""
Created on Fri Jul  9 01:21:56 2021

@author: santc
"""

import pandas as pd
import calendar
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from datetime import timedelta

print('\n\n########### CONFIGURAÇÕES ############')
print('\nUse a versão Python 3.x')
print('\nAs bibliotecas abaixo são necessárias (Instale com "pip install NOME_BIBLIOTECA"):')
print('\n   * pandas')
print('\n   * openpyxl')
print('\n   * calendar')
print('\n   * datetime')
print('\n\n########### INÍCIO ############')

nome_colaborador = input('\nDigite o seu nome: ')
cargo_colaborador = input('\nDigite o seu cargo: ')
fonte_csv = input('\nDigite o caminho do arquivo CSV (não é necessário a extensão) exportado pelo JIRA: ')
arquivo_xlsx = input('\nDigite o caminho para salvar o arquivo XLSX (não é necessário a extensão): ')
input_inicio = input('\nDigite a hora de início (hh:mm:ss): ')
input_fim = input('\nDigite a hora de fim (hh:mm:ss): ')
input_intervalo = input('\nDigite a hora de intervalo (hh:mm:ss): ')

ponto = pd.read_csv(fonte_csv +'.csv', sep = ',', encoding = 'utf-8')
ponto.head()
ponto.count()

def fatiarVerticalmenteDataFrame(coluna_inicial, coluna_final):
    return ponto.iloc[:,coluna_inicial:coluna_final]

data_anterior = ['']
def prepararTempo(data_str, hora, tempo_repetido):
    data = data_str
    if(data == data_anterior[0]):
        hora = tempo_repetido
        
    data_anterior[0] = data
    return hora

def criarColunaTempoStringComLinhasZeradasSeOrigemRepetir(coluna_origem,hora):
    return ponto[coluna_origem].apply(lambda data_str: prepararTempo(data_str[0:10], hora, '00:00:00'))

def criarColunaTimedeltaComMesmoTamanhoDaOrigem(origem):
    return [pd.Timedelta(h) for h in origem]

def criarDataframeQtdeHorasTimedelta(inicio_delta, fim_delta, intevalo_delta):
    qtde_horas_df = pd.DataFrame()
    qtde_horas_df = qtde_horas_df.assign(inicio_delta=inicio_delta)
    qtde_horas_df = qtde_horas_df.assign(fim_delta=fim_delta)
    qtde_horas_df = qtde_horas_df.assign(intevalo_delta=intevalo_delta)
    qtde_horas_delta = qtde_horas_df['fim_delta'] - qtde_horas_df['inicio_delta'] - qtde_horas_df['intevalo_delta']
    qtde_horas_df = qtde_horas_df.assign(qtde_horas=qtde_horas_delta)
    
    return qtde_horas_df

def converterSegundosParaStringHora(segundos):
    s = segundos
    hours, remainder = divmod(s, 3600)
    minutes, seconds = divmod(remainder, 60)
    return '{:02}:{:02}:{:02}'.format(int(hours), int(minutes), int(seconds))

def calcularHorasTotaisTimedelta(coluna, df):
    qtde_horas = df[coluna]
    return converterSegundosParaStringHora(qtde_horas.sum().total_seconds())

def converterColunaHoraTimedeltaParaHoraString(coluna, df):
    return df[coluna].apply(lambda qtde_horas: converterSegundosParaStringHora(qtde_horas.total_seconds()))

def percorrerTodasStringsDataEExcluirHora(coluna):
    return ponto[coluna].apply(lambda data_str: data_str[0:10])

def converterColunaStringDataParaTimestamp(coluna):
    return pd.to_datetime(ponto[coluna], format='%Y-%m-%d')

def getDado(linha,coluna):
    return ponto.loc[linha][coluna]

def criarListaTimestampDiasDoMes(numero_dias, mes, ano):
    return [pd.Timestamp(ano, mes, day) for day in range(1, numero_dias+1)]

def criarDataFrameDiasDoMes():
    ponto_primeiro_dia = getDado(0,'data de Trabalho')
    ano = ponto_primeiro_dia.year
    mes = ponto_primeiro_dia.month
    numero_dias = calendar.monthrange(ano, mes)[1]
    
    dias_lista = criarListaTimestampDiasDoMes(numero_dias, mes, ano)
    
    dias = {'data de Trabalho': dias_lista}
    return pd.DataFrame(dias)

def converterColunaTimestampParaStringData(coluna):
    return ponto_merge_dias[coluna].apply(lambda ts: ts.strftime('%d/%m/%Y'))

def converterFloatParaHora(tempo_float):
    horas_timedelta = pd.Timedelta(hours=tempo_float)
    segundos = horas_timedelta.total_seconds()

    return converterSegundosParaStringHora(segundos)
    

def getTotalHoras(coluna):
    total_horas = ponto_merge_dias[coluna].sum()
    return converterFloatParaHora(total_horas)

def timeDeltaHandler(timedelta):
    return pd.Timedelta('00:00:00') if pd.isnull(timedelta) else timedelta

def alterarNaTParaTimeDeltaZero(coluna):
    return ponto_merge_dias[coluna].apply(lambda timedelta: timeDeltaHandler(timedelta))

def converterColunaFloatParaStringData(coluna):
    return ponto_merge_dias['Horas'].apply(lambda tempo_float: converterFloatParaHora(tempo_float))

def getNomeMes(mes):
    meses = {
        1:'Janeiro',
        2:'Fevereiro',
        3:'Março',
        4:'Abril',
        5:'Maio',
        6:'Junho',
        7:'Julho',
        8:'Agosto',
        9:'Setembro',
        10:'Outubro',
        11:'Novembro',
        12:'Dezembro'}
    return meses[mes]

def converterCelulasParaFormato(coluna, formato):
    count = 0
    for item in ws[coluna]:
        if(count == 0):
            count += 1
            continue
        celula = coluna + str(item.row)
        ws[celula].number_format = formato

def aplicarFormulaEFormatarColunaQtdeHoras(coluna):
    count = 0
    for item in ws[coluna]:
        if(count == 0):
            count += 1
            continue
        celula = coluna + str(item.row)
        ws[celula] = '=(C'+str(item.row)+'-B'+str(item.row)+')-D'+str(item.row)
        ws[celula].number_format = 'hh:mm:ss'

def mesclarCelulas(coluna,proxima_coluna,celulas_merge_param):
    count = 0
    celula_selecionada = None
    celulas_merge = []
    
    if(len(celulas_merge_param) > 0):
        for celulas_merge in celulas_merge_param:
            ws.merge_cells(celulas_merge)
            
        return list(map(lambda celula: celula.replace(coluna,proxima_coluna), celulas_merge_param))
    else:
        for celula in ws[coluna]:
            # Pula o cabeçalho
            if(count == 0):
                count += 1
                continue
            
            if(pd.isnull(celula.value) == False and celula.value != '00:00:00'):
                if(pd.isnull(celula_selecionada) == False and celula_selecionada != '00:00:00'):
                    celula_merge = coluna + str(celula.row - 1)
                    celula_merge = celula_selecionada +':'+ celula_merge
                    celulas_merge.append(celula_merge)
                    ws.merge_cells(celula_merge)
                
                celula_selecionada = coluna + str(celula.row)
             
        celula_merge = coluna + str(celula.row)
        celula_merge = celula_selecionada +':'+ celula_merge
        celulas_merge.append(celula_merge)
        ws.merge_cells(celula_merge)
            
        return list(map(lambda celula: celula.replace(coluna,proxima_coluna), celulas_merge))

ponto = fatiarVerticalmenteDataFrame(0, 4)

inicio = criarColunaTempoStringComLinhasZeradasSeOrigemRepetir('data de Trabalho',input_inicio)
fim = criarColunaTempoStringComLinhasZeradasSeOrigemRepetir('data de Trabalho',input_fim)
intevalo = criarColunaTempoStringComLinhasZeradasSeOrigemRepetir('data de Trabalho',input_intervalo)

inicio_delta = criarColunaTimedeltaComMesmoTamanhoDaOrigem(inicio)
fim_delta = criarColunaTimedeltaComMesmoTamanhoDaOrigem(fim)
intevalo_delta = criarColunaTimedeltaComMesmoTamanhoDaOrigem(intevalo)

qtde_horas_df = criarDataframeQtdeHorasTimedelta(inicio_delta, fim_delta, intevalo_delta)

qtde_horas = converterColunaHoraTimedeltaParaHoraString('qtde_horas', qtde_horas_df)

ponto['data de Trabalho'] = percorrerTodasStringsDataEExcluirHora('data de Trabalho')
ponto['data de Trabalho'] = converterColunaStringDataParaTimestamp('data de Trabalho')

ponto = ponto.assign(inicio=inicio)
ponto = ponto.assign(fim=fim)
ponto = ponto.assign(intervalo=intevalo)
ponto = ponto.assign(qtde_horas=qtde_horas)

ponto_primeiro_dia = getDado(0,'data de Trabalho')
dias_do_mes = criarDataFrameDiasDoMes()

# Merge entre os dataframes - Foi criado o dataframe auxiliar dias para
# juntar com a planilha de horas trabalhadas (ponto) de modo que ela possua
# todos os dias do mês, inclusive os dias de folga.
ponto_merge_dias = pd.merge(ponto, dias_do_mes, how='right')

ponto_merge_dias['Horas'] = ponto_merge_dias['Horas'].fillna(0)
ponto_merge_dias['Questão-chave'] = ponto_merge_dias['Questão-chave'].fillna('')
ponto_merge_dias['Emissão de resumo'] = ponto_merge_dias['Emissão de resumo'].fillna('')

ponto_merge_dias['inicio'] = alterarNaTParaTimeDeltaZero('inicio')
ponto_merge_dias['fim'] = alterarNaTParaTimeDeltaZero('fim')
ponto_merge_dias['intervalo'] = alterarNaTParaTimeDeltaZero('intervalo')
ponto_merge_dias['qtde_horas'] = alterarNaTParaTimeDeltaZero('qtde_horas')

ponto_merge_dias['data de Trabalho'] = converterColunaTimestampParaStringData('data de Trabalho')

ponto_merge_dias['Horas']

ponto_merge_dias['Horas'] = converterColunaFloatParaStringData('Horas')

ponto_merge_dias = ponto_merge_dias[['data de Trabalho','inicio','fim','intervalo',
                                     'qtde_horas','Questão-chave','Horas','Emissão de resumo']]



def handlerHoraTimedeltaOuHoraString(hora, tipo_retorno):
    if(isinstance(hora, str) and hora == '00:00:00'):
        return tipo_retorno
    elif(isinstance(hora, str) and hora != '00:00:00'):
        return hora
    elif(isinstance(hora, pd.Timedelta) and hora.value == 0):
        return tipo_retorno
    elif(isinstance(hora, pd.Timedelta) and hora.value != 0):
        return converterSegundosParaStringHora(hora.total_seconds())

ponto_merge_dias['inicio'] = ponto_merge_dias['inicio'].apply(lambda hora: handlerHoraTimedeltaOuHoraString(hora,''))
ponto_merge_dias['fim'] = ponto_merge_dias['fim'].apply(lambda hora: handlerHoraTimedeltaOuHoraString(hora,''))
ponto_merge_dias['intervalo'] = ponto_merge_dias['intervalo'].apply(lambda hora: handlerHoraTimedeltaOuHoraString(hora,'00:00:00'))
ponto_merge_dias['qtde_horas'] = ponto_merge_dias['qtde_horas'].apply(lambda hora: handlerHoraTimedeltaOuHoraString(hora,'00:00:00'))

ponto_merge_dias['data de Trabalho'] = ponto_merge_dias['data de Trabalho'].apply(lambda data_str: prepararTempo(data_str[0:11], data_str[0:11], ''))

ponto_merge_dias = ponto_merge_dias.rename(
    columns={
        'data de Trabalho': 'Data',
        'inicio': 'Início',
        'fim': 'Fim',
        'intervalo': 'Intervalo',
        'qtde_horas': 'Qtde Horas',
        'Questão-chave': 'Chamado',
        'Horas': 'Tempo Gasto',
        'Emissão de resumo': 'Atividades'})

ponto_merge_dias['Tempo Gasto'] = ponto_merge_dias['Tempo Gasto'].apply(lambda hora: timedelta(hours=int(hora[0:2]), minutes=int(hora[3:5]), seconds=int(hora[6:8])))

ponto_merge_dias.to_excel(arquivo_xlsx +'.xlsx')

wb = load_workbook(filename = arquivo_xlsx +'.xlsx')

ws = wb.active

ws.delete_cols(1)

# Título da planilha
ws.title = getNomeMes(ponto_primeiro_dia.month)

ws.column_dimensions["A"].width = 13
ws.column_dimensions["B"].width = 13
ws.column_dimensions["C"].width = 13
ws.column_dimensions["D"].width = 13
ws.column_dimensions["E"].width = 13
ws.column_dimensions["F"].width = 13
ws.column_dimensions["G"].width = 13
ws.column_dimensions["H"].width = 60

aplicarFormulaEFormatarColunaQtdeHoras('E')
converterCelulasParaFormato('G', 'hh:mm:ss')

qtde_horas_totais = WriteOnlyCell(ws, value='=SUM(B2:B'+ str(ws.max_row) +')')
tempo_gasto_total = '=SUM(C2:C'+ str(ws.max_row) +')'
ws.append(['',qtde_horas_totais,tempo_gasto_total])

qtde_horas_totais = 'B'+str(ws.max_row)
ws.move_range(qtde_horas_totais, cols=3, translate=True)

tempo_gasto_total = 'C'+str(ws.max_row)
ws.move_range(tempo_gasto_total, cols=4, translate=True)

qtde_horas_totais = 'E'+str(ws.max_row)
ws[qtde_horas_totais].number_format = '[h]:mm:ss'

tempo_gasto_total = 'G'+str(ws.max_row)
ws[tempo_gasto_total].number_format = '[h]:mm:ss'

todos_celulas = 'A1:H'+str(ws.max_row)
ws.move_range(todos_celulas, rows=4, cols=0, translate=True)

celulas_merge = mesclarCelulas('A','B',[])
celulas_merge = mesclarCelulas('B','C',celulas_merge)
celulas_merge = mesclarCelulas('C','D',celulas_merge)
celulas_merge = mesclarCelulas('D','E',celulas_merge)
mesclarCelulas('E','F',celulas_merge)

ws['A1'] = 'Horas Trabalhadas em '+ getNomeMes(ponto_primeiro_dia.month) +'/'+ str(ponto_primeiro_dia.year)
ws['A3'] = 'Profissional: '+ nome_colaborador +' ('+ cargo_colaborador +') Projeto: PD CASE TCE/SC outsourcing'

ws.merge_cells('A1:H1')
ws.merge_cells('A2:H2')
ws.merge_cells('A3:H3')
ws.merge_cells('A4:H4')

from openpyxl.styles import PatternFill, Border, Side, Alignment

fill = PatternFill("solid", fgColor="00EBF1DE")

border = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin',color='00000000'),
    top=Side(border_style='thin',color='00000000'),
    bottom=Side(border_style='thin',color='00000000'))

alignment = Alignment(
    horizontal='center', 
    vertical='center', 
    text_rotation=0, 
    wrap_text=False, 
    shrink_to_fit=False, 
    indent=0)


ws['A1'].fill = fill
ws['A3'].fill = fill
for coluna in range(1,9): ws.cell(row=5, column=coluna).fill = fill

for coluna in range(1,9): ws.cell(row=ws.max_row, column=coluna).fill = fill

for linha in range(1, ws.max_row + 1):
    for coluna in range(1, 9):
        ws.cell(row=linha, column=coluna).border = border
        
for linha in range(1, ws.max_row + 1):
    for coluna in range(1, 8):
        ws.cell(row=linha, column=coluna).alignment = alignment

alignment = Alignment(
    horizontal='right', 
    vertical='center', 
    text_rotation=0, 
    wrap_text=False, 
    shrink_to_fit=False, 
    indent=0)

total = 'D'+ str(ws.max_row)
ws[total] = 'Total:'

for coluna in range(1,5): ws.cell(row=ws.max_row, column=coluna).alignment = alignment

border = Border(
    left=Side(border_style=None),
    right=Side(border_style=None),
    top=Side(border_style='thin',color='00000000'),
    bottom=Side(border_style='thin',color='00000000'))

for coluna in range(1,5): ws.cell(row=ws.max_row, column=coluna).border = border

wb.save(filename = arquivo_xlsx +'.xlsx')